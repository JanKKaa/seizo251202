from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import UserPassesTestMixin
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.http import HttpResponse, JsonResponse
from django.core.mail import send_mail
from django.conf import settings
from .models import FaxStatus
from django.utils import timezone

from .models import MonAn, NhanVien, Holiday, Order
from .forms import MonAnForm  # Dùng form gốc
from collections import defaultdict
from calendar import monthrange
from datetime import date, datetime, timedelta, time
import csv
import json
from weasyprint import HTML
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ----------------- CẤU HÌNH -----------------
PAST_LOCK_ENABLED = True          # Khóa dữ liệu quá khứ
LOCK_EDIT_PAST_QUANTITY = True
NEXT_DAY_CUTOFF_HOUR = 13         # Sau giờ này không đặt được cho ngày mai
DISALLOW_SAME_DAY = True          # Không cho đặt cho chính ngày hôm nay
# --------------------------------------------

# ================== HELPERS =================
def today():
    return date.today()

def is_past(d: date, request=None):
    """
    Trả về True (tức KHÔNG cho đặt / sửa) nếu:
      - d < hôm nay (quá khứ)
      - DISALLOW_SAME_DAY và d == hôm nay (không cho đặt trong ngày)
      - d == ngày mai và đã qua giờ cắt (NEXT_DAY_CUTOFF_HOUR)
    Nếu đăng nhập bằng user 'kanri' thì luôn trả về False (không giới hạn).
    """
    # Nếu request có user là kanri thì không giới hạn
    if request and hasattr(request, "user") and request.user.is_authenticated and request.user.username == 'kanri':
        return False
    if not PAST_LOCK_ENABLED:
        return False
    today_d = today()
    if d < today_d:
        return True
    if DISALLOW_SAME_DAY and d == today_d:
        return True
    # Giới hạn đặt cho ngày mai sau 13:00
    if d == today_d + timedelta(days=1):
        now = datetime.now()
        if now.time() >= time(NEXT_DAY_CUTOFF_HOUR, 0):
            return True
    return False

def current_period(ref: date = None):
    """
    Kỳ: 16 (tháng N) -> 15 (tháng N+1)
    Trả về: (start_date, end_date, period_year, period_month)
    period_year / period_month là tháng phần 16 thuộc về.
    """
    if ref is None:
        ref = today()
    if ref.day >= 16:
        start = date(ref.year, ref.month, 16)
        if ref.month == 12:
            end = date(ref.year + 1, 1, 15)
        else:
            end = date(ref.year, ref.month + 1, 15)
        return start, end, ref.year, ref.month
    else:
        if ref.month == 1:
            start = date(ref.year - 1, 12, 16)
            end = date(ref.year, 1, 15)
            return start, end, ref.year - 1, 12
        else:
            start = date(ref.year, ref.month - 1, 16)
            end = date(ref.year, ref.month, 15)
            return start, end, ref.year, ref.month - 1

def working_days(start_d: date, end_d: date, holidays: set):
    d = start_d
    while d <= end_d:
        if d.weekday() < 5 and d not in holidays:
            yield d
        d += timedelta(days=1)

# ================== AUTH / MENU =============
class OnlyKanriMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.username == 'kanri'

class MenuListView(ListView):
    model = MonAn
    template_name = 'menu/menu_list.html'
    context_object_name = 'object_list'

    def dispatch(self, request, *args, **kwargs):
        ma_nv = request.session.get('ma_nv')
        if not ma_nv:
            return redirect('menu:dangnhap')
        nv = NhanVien.objects.filter(ma_so=ma_nv).first()
        self.extra_context = {
            'ma_nv': ma_nv,
            'ten_nv': nv.ten if nv else ""
        }
        return super().dispatch(request, *args, **kwargs)

class MenuCreateView(OnlyKanriMixin, CreateView):
    model = MonAn
    form_class = MonAnForm
    template_name = 'menu/menu_form.html'
    success_url = reverse_lazy('menu:list')

class MenuUpdateView(OnlyKanriMixin, UpdateView):
    model = MonAn
    form_class = MonAnForm
    template_name = 'menu/menu_form.html'
    success_url = reverse_lazy('menu:list')

class MenuDeleteView(OnlyKanriMixin, DeleteView):
    model = MonAn
    template_name = 'menu/menu_confirm_delete.html'
    success_url = reverse_lazy('menu:list')

def dangnhap_ma_nv(request):
    error = None
    if request.method == 'POST':
        ma_nv = request.POST.get('ma_nv', '').strip()
        if not ma_nv:
            error = "社員番号を入力してください。"
        else:
            nv = NhanVien.objects.filter(ma_so=ma_nv).first()
            if not nv:
                error = "この社員番号は存在しません。管理課によって登録されていません。"
            else:
                request.session['ma_nv'] = ma_nv
                return redirect('menu:order_history')
    return render(request, 'menu/dangnhap.html', {'error': error})

# ================== ORDER PLACEMENT =========
def order_menu(request, pk):
    mon = get_object_or_404(MonAn, pk=pk)
    ma_nv = request.session.get('ma_nv')
    if not ma_nv:
        return redirect('menu:dangnhap')
    nv = NhanVien.objects.filter(ma_so=ma_nv).first()
    ten_nv = nv.ten if nv else ""
    error = None

    cur_start, cur_end, cur_year, cur_month = current_period()

    if request.method == 'POST':
        holidays = set(Holiday.objects.values_list('date', flat=True))
        calamviec = request.POST.get('calamviec', '日勤')

        # Account đặc biệt chọn ngày lẻ
        if ma_nv == "1000":
            order_dates_str = request.POST.get('order_dates', '')
            date_list = [s for s in order_dates_str.split(',') if s.strip()]
            try:
                so_luong = int(request.POST.get('so_luong', 1))
                if so_luong < 1:
                    so_luong = 1
            except:
                so_luong = 1
            created = 0
            skipped_past = 0
            first_year = first_month = None
            for raw in date_list:
                try:
                    d_sel = datetime.strptime(raw.strip(), "%Y-%m-%d").date()
                except:
                    continue
                if is_past(d_sel, request):
                    skipped_past += 1
                    continue
                if d_sel.weekday() >= 5 or d_sel in holidays:
                    continue
                if first_year is None:
                    first_year = d_sel.year
                    first_month = d_sel.month
                # Không ghi đè nếu đã có (giữ lịch sử)
                Order.objects.get_or_create(
                    ma_nv=ma_nv,
                    ngay_giao=d_sel,
                    defaults={
                        'ten_nv': ten_nv,
                        'mon_an': mon,
                        'so_luong': so_luong,
                        'ghi_chu': '',
                        'calamviec': calamviec
                    }
                )
                created += 1
            if created:
                send_mail(
                    subject='【注文内容変更通知】',
                    message=f'社員番号: {ma_nv}\n氏名: {ten_nv}\nお弁当: {mon.ten}\n年月: {first_year}年{first_month}月\nが新規登録または変更されました。',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=['kanri_3@hayashi-p.co.jp'],
                    fail_silently=True,
                )
                if 'ma_nv' in request.session:
                    del request.session['ma_nv']
                return render(request, 'menu/order_success.html', {
                    'mon': mon, 'ma_nv': ma_nv, 'ten_nv': ten_nv,
                    'skipped_past': skipped_past
                })
            else:
                error = error or "選択した日に注文できません。"
        else:
            order_year_str = request.POST.get('order_year', '').strip()
            order_month_str = request.POST.get('order_month', '').strip()
            if not order_year_str or not order_month_str:
                error = "注文月を選択してください。"
            else:
                order_year = int(order_year_str)
                order_month = int(order_month_str)
                so_luong = 1  # Chỉ 1000 mới thay đổi

                # Kỳ chọn
                start_date = date(order_year, order_month, 16)
                if order_month == 12:
                    end_date = date(order_year + 1, 1, 15)
                else:
                    end_date = date(order_year, order_month + 1, 15)

                if end_date < today():
                    error = "過去の期間には注文できません。"
                else:
                    # Không cho đổi món trong kỳ hiện tại nếu đã có
                    if order_year == cur_year and order_month == cur_month:
                        exist = Order.objects.filter(
                            ma_nv=ma_nv,
                            ngay_giao__gte=start_date,
                            ngay_giao__lte=end_date
                        )
                        if exist.exists():
                            mon_cu = exist.first().mon_an
                            if mon_cu.id != mon.id:
                                error = f"{order_year}年{order_month}月分（{start_date.strftime('%Y/%m/%d')}～{end_date.strftime('%Y/%m/%d')}）は既に「{mon_cu.ten}」を注文しています。"

                if not error:
                    created = 0
                    skipped_past = 0
                    for d in working_days(start_date, end_date, holidays):
                        if is_past(d, request):
                            skipped_past += 1
                            continue
                        # Không ghi đè ngày đã có (giữ lựa chọn trước đó)
                        Order.objects.get_or_create(
                            ma_nv=ma_nv,
                            ngay_giao=d,
                            defaults={
                                'ten_nv': ten_nv,
                                'mon_an': mon,
                                'so_luong': so_luong,
                                'ghi_chu': '',
                                'calamviec': calamviec
                            }
                        )
                        created += 1
                    if created:
                        send_mail(
                            subject='【注文内容変更通知】',
                            message=f'社員番号: {ma_nv}\n氏名: {ten_nv}\nお弁当: {mon.ten}\n年月: {order_year}年{order_month}月\nが新規登録または変更されました。',
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=['kanri_3@hayashi-p.co.jp'],
                            fail_silently=True,
                        )
                        if 'ma_nv' in request.session:
                            del request.session['ma_nv']
                        return render(request, 'menu/order_success.html', {
                            'mon': mon, 'ma_nv': ma_nv, 'ten_nv': ten_nv,
                            'skipped_past': skipped_past
                        })
                    else:
                        error = error or "この月には注文可能な平日がありません。"

    # Context hiển thị
    holidays = Holiday.objects.values_list('date', flat=True)
    holidays_str = [d.strftime('%Y-%m-%d') for d in holidays]
    orders = Order.objects.filter(ma_nv=ma_nv)
    mon_by_month = {}
    for o in orders:
        if o.ngay_giao.day >= 16:
            key = f"{o.ngay_giao.year}-{o.ngay_giao.month}"
            mon_by_month.setdefault(key, o.mon_an.ten)
    da_dat_mon = any(o.mon_an.id == mon.id for o in orders)

    return render(request, 'menu/order_menu.html', {
        'mon': mon,
        'ma_nv': ma_nv,
        'ten_nv': ten_nv,
        'error': error,
        'holidays': holidays_str,
        'mon_by_month_json': json.dumps(mon_by_month, ensure_ascii=False),
        'da_dat_mon': da_dat_mon,
        'allow_so_luong': ma_nv == "1000",
        'data_day': [],
        'data_night': [],
        'data_koujika': [],
        'data_kyakusama': [],
        'days': [],
    })

# =============== NHÂN VIÊN =================
class NhanVienForm(forms.ModelForm):
    class Meta:
        model = NhanVien
        fields = ['ma_so', 'ten']
        labels = {'ma_so': '社員番号', 'ten': '氏名'}

def nhanvien_list(request):
    lst = sorted(NhanVien.objects.all(), key=lambda nv: int(nv.ma_so))
    return render(request, 'menu/nhanvien.html', {'nhanvien_list': lst})

def nhanvien_create(request):
    if request.method == 'POST':
        form = NhanVienForm(request.POST)
        if form.is_valid():
            if NhanVien.objects.filter(ma_so=form.cleaned_data['ma_so']).exists():
                form.add_error('ma_so', 'この社員番号は既に存在します。')
            else:
                form.save()
                messages.success(request, '社員を追加しました。')
                return redirect('menu:nhanvien_list')
    else:
        form = NhanVienForm()
    return render(request, 'menu/nhanvien_form.html', {'form': form, 'title': '新しい社員を追加'})

def nhanvien_update(request, pk):
    if not (request.user.is_authenticated and request.user.username == 'kanri'):
        raise PermissionDenied
    nv = get_object_or_404(NhanVien, pk=pk)
    if request.method == 'POST':
        form = NhanVienForm(request.POST, instance=nv)
        if form.is_valid():
            form.save()
            return redirect('menu:nhanvien_list')
    else:
        form = NhanVienForm(instance=nv)
    return render(request, 'menu/nhanvien_form.html', {'form': form, 'title': '社員情報を編集'})

def nhanvien_delete(request, pk):
    if not (request.user.is_authenticated and request.user.username == 'kanri'):
        raise PermissionDenied
    nv = get_object_or_404(NhanVien, pk=pk)
    if request.method == 'POST':
        nv.delete()
        return redirect('menu:nhanvien_list')
    return render(request, 'menu/nhanvien_confirm_delete.html', {'nhanvien': nv})

def custom_permission_denied_view(request, exception=None):
    return render(request, 'menu/403.html', status=403)

@csrf_exempt
def logout_nv(request):
    request.session.pop('ma_nv', None)
    return redirect('menu:dangnhap')

# =============== HOLIDAYS ===================
def holiday_list(request):
    if not request.user.is_authenticated or request.user.username != 'kanri':
        return redirect('menu:list')
    error = message = None
    if request.method == 'POST' and 'holiday_date' in request.POST:
        dates_raw = request.POST.get('holiday_date', '')
        note = request.POST.get('holiday_note', '')
        created = existed = 0
        for s in dates_raw.split(','):
            s = s.strip()
            if not s:
                continue
            try:
                d_obj = datetime.strptime(s, "%Y-%m-%d").date()
            except:
                continue
            if not Holiday.objects.filter(date=d_obj).exists():
                Holiday.objects.create(date=d_obj, note=note)
                Order.objects.filter(ngay_giao=d_obj).delete()
                created += 1
            else:
                existed += 1
        if created:
            message = f"{created}件の休日を登録しました。"
        if existed:
            error = f"{existed}件の選択した日付はすでに登録されています。"
    holidays = Holiday.objects.all().order_by('date')
    return render(request, 'menu/holiday_list.html', {
        'holidays': holidays,
        'error': error,
        'message': message
    })

@require_POST
def holiday_delete(request, pk):
    if not request.user.is_authenticated or request.user.username != 'kanri':
        return redirect('menu:list')
    h = get_object_or_404(Holiday, pk=pk)
    h.delete()
    messages.success(request, "休日を削除しました。")
    return redirect('menu:holiday_list')

# =============== ORDER HISTORY / EDIT =======
def order_history(request):
    ma_nv = request.session.get('ma_nv')
    if not ma_nv:
        return redirect('menu:dangnhap')
    orders = Order.objects.filter(ma_nv=ma_nv).select_related('mon_an').order_by('ngay_giao')
    nv = NhanVien.objects.filter(ma_so=ma_nv).first()
    ten_nv = nv.ten if nv else ""
    calamviec = orders.first().calamviec if orders.exists() else None
    holidays = [d.strftime('%Y-%m-%d') for d in Holiday.objects.values_list('date', flat=True)]
    is_kanri = request.user.is_authenticated and request.user.username == "kanri"

    # Xử lý thay đổi ca làm việc từng ngày
    if is_kanri and request.method == 'POST':
        for order in orders:
            calamviec_new = request.POST.get(f'calamviec_{order.pk}', order.calamviec)
            if calamviec_new != order.calamviec:
                order.calamviec = calamviec_new
                order.save()
        messages.success(request, "勤務区分を更新しました。")
        return redirect('menu:order_history')

    return render(request, 'menu/order_history.html', {
        'orders': orders,
        'ma_nv': ma_nv,
        'ten_nv': ten_nv,
        'holidays': holidays,
        'calamviec': calamviec,
        'is_kanri': is_kanri,
    })

def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)
    ma_nv = request.session.get('ma_nv')
    if not ma_nv or order.ma_nv != ma_nv:
        return redirect('menu:order_history')
    if is_past(order.ngay_giao, request):
        messages.error(request, "過去の日付は修正できません。")
        return redirect('menu:order_history')
    error = None
    if request.method == 'POST':
        so_luong = request.POST.get('so_luong', 1)
        ghi_chu = request.POST.get('ghi_chu', '')
        ngay_giao_str = request.POST.get('ngay_giao', '')
        try:
            ngay_giao = datetime.strptime(ngay_giao_str, "%Y-%m-%d").date()
        except:
            error = "配達日の形式が正しくありません。"
        else:
            if is_past(ngay_giao, request):
                error = "過去の日付へ変更できません。"
            else:
                order.so_luong = so_luong
                order.ghi_chu = ghi_chu
                order.ngay_giao = ngay_giao
                order.save()
                return redirect('menu:order_history')
    return render(request, 'menu/order_edit.html', {'order': order, 'error': error})

def order_delete(request, pk):
    order = get_object_or_404(Order, pk=pk)
    ma_nv = request.session.get('ma_nv')
    is_kanri = request.user.is_authenticated and request.user.username == 'kanri'
    
    if not ma_nv or (order.ma_nv != ma_nv and not is_kanri):
        return redirect('menu:order_history')
    
    # Chỉ kiểm tra is_past nếu không phải kanri
    if not is_kanri and is_past(order.ngay_giao, request):
        messages.error(request, "過去の日付の注文はキャンセルできません。")
        return redirect('menu:order_history')
    
    email_to = 'kanri_3@hayashi-p.co.jp'
    order.delete()
    send_mail(
        subject='【注文キャンセル通知】',
        message=f'社員番号: {order.ma_nv}\n氏名: {order.ten_nv}\n日付: {order.ngay_giao}\n料理: {order.mon_an.ten}\n数量: {order.so_luong}\nがキャンセルされました。',
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[email_to],
        fail_silently=True,
    )
    return redirect('menu:order_history')

def order_reorder(request, pk):
    """Đặt lại order quá khứ - có thể đặt cho bất kỳ ngày nào (kể cả quá khứ)"""
    order = get_object_or_404(Order, pk=pk)
    ma_nv = request.session.get('ma_nv')
    if not ma_nv:
        return redirect('menu:dangnhap')
    
    nv = NhanVien.objects.filter(ma_so=ma_nv).first()
    ten_nv = nv.ten if nv else ""
    error = None
    
    if request.method == 'POST':
        so_luong = request.POST.get('so_luong', 1)
        ghi_chu = request.POST.get('ghi_chu', '')
        ngay_giao_str = request.POST.get('ngay_giao', '')
        calamviec = request.POST.get('calamviec', order.calamviec or '日勤')
        
        try:
            so_luong = int(so_luong)
            if so_luong < 1:
                so_luong = 1
        except (ValueError, TypeError):
            so_luong = 1
        
        try:
            ngay_giao = datetime.strptime(ngay_giao_str, "%Y-%m-%d").date()
        except ValueError:
            error = "配達日の形式が正しくありません。"
        else:
            # Kiểm tra ngày là weekday và không phải holiday
            holidays = set(Holiday.objects.values_list('date', flat=True))
            if ngay_giao.weekday() >= 5:
                error = "土日祝日には注文できません。"
            elif ngay_giao in holidays:
                error = "この日は休日です。注文できません。"
            else:
                # Cập nhật order (cho phép cả quá khứ)
                order.so_luong = so_luong
                order.ghi_chu = ghi_chu
                order.ngay_giao = ngay_giao
                order.calamviec = calamviec
                order.save()
                
                # Gửi email thông báo
                send_mail(
                    subject='【注文内容変更通知】',
                    message=f'社員番号: {ma_nv}\n氏名: {ten_nv}\nお弁当: {order.mon_an.ten}\n配達日: {ngay_giao}\n数量: {so_luong}\n勤務区分: {calamviec}\nが変更されました。',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=['kanri_3@hayashi-p.co.jp'],
                    fail_silently=True,
                )
                messages.success(request, "注文を再登録しました。")
                return redirect('menu:order_history')
    
    holidays = [d.strftime('%Y-%m-%d') for d in Holiday.objects.values_list('date', flat=True)]
    
    return render(request, 'menu/order_reorder.html', {
        'order': order,
        'ma_nv': ma_nv,
        'ten_nv': ten_nv,
        'error': error,
        'holidays': holidays,
    })

def order_delete_all(request, ma_nv):
    # Chỉ xóa đơn tương lai
    Order.objects.filter(ma_nv=ma_nv, ngay_giao__gte=today()).delete()
    return redirect('menu:order_history')

# =============== ADMIN DASHBOARD =============
def order_kanri(request):
    t = today()
    month_param = request.GET.get('month')
    if month_param:
        y, m = map(int, month_param.split('-'))
        start_date = date(y, m, 16)
        end_date = date(y + 1, 1, 15) if m == 12 else date(y, m + 1, 15)
        period_year, period_month = y, m
    else:
        start_date, end_date, period_year, period_month = current_period(t)

    days = []
    d = start_date
    while d <= end_date:
        days.append(d)
        d += timedelta(days=1)

    nhanviens = NhanVien.objects.all()
    orders = Order.objects.filter(ngay_giao__gte=start_date, ngay_giao__lte=end_date).select_related('mon_an')
    orders_dict = {(o.ma_nv, o.ngay_giao): o for o in orders}

    groups = {
        'day': [],
        'night': [],
        'koujika': [],
        'kyakusama': []
    }

    for nv in nhanviens:
        # Tìm tất cả loại ca làm việc của nhân viên trong kỳ
        calamviec_set = set()
        orders_in_period = []
        for d2 in days:
            o = orders_dict.get((nv.ma_so, d2))
            if o:
                orders_in_period.append(o)
                calamviec_set.add(o.calamviec)
        for calamviec in calamviec_set or ['日勤']:
            # Lọc các order thuộc loại ca này
            orders_calam = [o for o in orders_in_period if o.calamviec == calamviec]
            if not orders_calam:
                continue
            latest_order = max(orders_calam, key=lambda o: o.ngay_giao)
            row = {
                'ma_so': nv.ma_so,
                'ten': nv.ten,
                'mon_an': latest_order.mon_an.ten if latest_order else '',
                'gia2': int(latest_order.mon_an.gia2) if latest_order else 0,
                'orders': [],
                'calamviec': calamviec,
            }
            total_qty = 0
            for d2 in days:
                o = orders_dict.get((nv.ma_so, d2))
                if o and o.calamviec == calamviec:
                    row['orders'].append(o.so_luong)
                    total_qty += o.so_luong
                else:
                    row['orders'].append(0)
            row['so_luong'] = total_qty
            row['total'] = row['gia2'] * total_qty
            row['orders_with_day'] = list(zip(row['orders'], days))
            if calamviec == '夜勤':
                groups['night'].append(row)
            elif calamviec == '工作課':
                groups['koujika'].append(row)
            elif calamviec == '客様':
                groups['kyakusama'].append(row)
            else:
                groups['day'].append(row)

    # --- Sắp xếp summary_list theo thứ tự ưu tiên ---
    PRIORITY_NAMES = [
        "椿弁当",
        "椿弁当(半ライス)",
        "椿弁当(おかずのみ)",
        "はなまる弁当(半ライス)",
        "はなまる弁当(おかずのみ)",
    ]

    # Lấy giá và tổng tiền theo giá2 nếu có
    summary = defaultdict(lambda: {'so_luong': 0, 'tong_tien': 0, 'gia': 0, 'gia2': None, 'tong_tien2': 0})
    for o in orders:
        ten_mon = o.mon_an.ten
        s = summary[ten_mon]
        s['so_luong'] += o.so_luong
        s['gia'] = o.mon_an.gia
        s['gia2'] = o.mon_an.gia2
        s['tong_tien'] += o.so_luong * o.mon_an.gia
        if o.mon_an.gia2:
            s['tong_tien2'] += o.so_luong * o.mon_an.gia2

    summary_list = [
        {
            'ten': k,
            'so_luong': v['so_luong'],
            'gia': v['gia'],
            'gia2': v['gia2'],
            'tong_tien': v['tong_tien'],
            'tong_tien2': v['tong_tien2'] if v['gia2'] else None,
        }
        for k, v in summary.items()
    ]

    # Hàm sắp xếp theo thứ tự ưu tiên
    def sort_key(mon):
        try:
            return (PRIORITY_NAMES.index(mon['ten']),)
        except ValueError:
            return (len(PRIORITY_NAMES), mon['ten'])

    summary_list = sorted(summary_list, key=sort_key)

    orders_by_day = {}
    for d2 in days:
        od = [o for o in orders if o.ngay_giao == d2]
        orders_by_day[d2] = {
            'count': len(od),
            'nguoi_dat': [o.ten_nv for o in od],
            'mon_an': [o.mon_an.ten for o in od],
            'tong_tien': sum(o.mon_an.gia * o.so_luong for o in od)
        }

    for key in groups:
        groups[key].sort(key=lambda x: int(x['ma_so']))

    # Sắp xếp từng nhóm theo thứ tự tên món ăn ưu tiên
    def row_sort_key(row):
        try:
            return (PRIORITY_NAMES.index(row['mon_an']), row['mon_an'])
        except ValueError:
            return (len(PRIORITY_NAMES), row['mon_an'])

    groups['day'].sort(key=row_sort_key)
    groups['night'].sort(key=row_sort_key)
    groups['koujika'].sort(key=row_sort_key)
    groups['kyakusama'].sort(key=row_sort_key)

    return render(request, 'menu/order_kanri.html', {
        'days': days,
        'data_day': groups['day'],
        'data_night': groups['night'],
        'data_koujika': groups['koujika'],
        'data_kyakusama': groups['kyakusama'],
        'year': period_year,
        'month': period_month,
        'orders_by_day': orders_by_day,
        'summary_list': summary_list,
        'ky_start': start_date,
        'ky_end': end_date,
    })

def order_kanri_csv(request):
    t = today()
    month_str = request.GET.get('month')
    if month_str:
        year, month = map(int, month_str.split('-'))
    else:
        year, month = t.year, t.month
    days_in_month = monthrange(year, month)[1]
    days = [date(year, month, d) for d in range(1, days_in_month + 1)]
    nhanviens = NhanVien.objects.all()
    orders = Order.objects.filter(ngay_giao__year=year, ngay_giao__month=month).select_related('mon_an')
    orders_dict = {(o.ma_nv, o.ngay_giao): o for o in orders}

    # --- Thứ tự ưu tiên tên món ăn ---
    PRIORITY_NAMES = [
        "椿弁当",
        "椿弁当(半ライス)",
        "椿弁当(おかずのみ)",
        "はなまる弁当(半ライス)",
        "はなまる弁当(おかずのみ)",
    ]

    data = []
    for nv in nhanviens:
        row = {'ma_so': nv.ma_so, 'ten': nv.ten, 'mon_an': '', 'gia': 0, 'orders': []}
        for d2 in days:
            o = orders_dict.get((nv.ma_so, d2))
            if o and not row['mon_an']:
                row['mon_an'] = o.mon_an.ten
                row['gia'] = int(o.mon_an.gia)
            row['orders'].append(o is not None)
        so_ngay = sum(1 for x in row['orders'] if x)
        row['total'] = row['gia'] * so_ngay
        # Ẩn nhân viên không có order nào
        if so_ngay > 0:
            data.append(row)

    # Sắp xếp theo thứ tự tên món ăn ưu tiên
    def row_sort_key(row):
        try:
            return (PRIORITY_NAMES.index(row['mon_an']), row['mon_an'])
        except ValueError:
            return (len(PRIORITY_NAMES), row['mon_an'])

    data.sort(key=row_sort_key)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="order_kanri.csv"'
    writer = csv.writer(response)
    header = ['社員番号', '氏名', '料理名'] + [str(d.day) for d in days] + ['合計金額']
    writer.writerow(header)
    for r in data:
        line = [r['ma_so'], r['ten'], r['mon_an']]
        line.extend('済' if b else '未注文' for b in r['orders'])
        line.append(r['total'])
        writer.writerow(line)
    return response

def order_kanri_pdf(request):
    t = today()
    month_str = request.GET.get('month')
    if month_str:
        year, month = map(int, month_str.split('-'))
    else:
        year, month = t.year, t.month
    start_date = date(year, month, 16)
    end_date = date(year + 1, 1, 15) if month == 12 else date(year, month + 1, 15)
    days = []
    d = start_date
    while d <= end_date:
        days.append(d)
        d += timedelta(days=1)

    nhanviens = NhanVien.objects.all()
    orders = Order.objects.filter(ngay_giao__gte=start_date, ngay_giao__lte=end_date).select_related('mon_an')
    orders_dict = {(o.ma_nv, o.ngay_giao): o for o in orders}

    groups = {'day': [], 'night': [], 'koujika': [], 'kyakusama': []}
    for nv in nhanviens:
        row = {'ma_so': nv.ma_so, 'ten': nv.ten, 'mon_an': '', 'gia': 0, 'orders': [], 'calamviec': '日勤'}
        calam = None
        total_qty = 0
        for d2 in days:
            o = orders_dict.get((nv.ma_so, d2))
            if o:
                if not row['mon_an']:
                    row['mon_an'] = o.mon_an.ten
                    row['gia'] = int(o.mon_an.gia2)
                if not calam:
                    calam = o.calamviec
                row['orders'].append(o.so_luong)
                total_qty += o.so_luong
            else:
                row['orders'].append(0)
        row['calamviec'] = calam or '日勤'
        row['so_luong'] = total_qty
        row['total'] = row['gia2'] * total_qty
        row['orders_with_day'] = list(zip(row['orders'], days))
        # Ẩn nhân viên không có order nào
        if total_qty == 0:
            continue
        if row['calamviec'] == '夜勤':
            groups['night'].append(row)
        elif row['calamviec'] == '工作課':
            groups['koujika'].append(row)
        elif row['calamviec'] == '客様':
            groups['kyakusama'].append(row)
        else:
            groups['day'].append(row)

    # --- Thứ tự ưu tiên tên món ăn ---
    PRIORITY_NAMES = [
        "椿弁当",
        "椿弁当(半ライス)",
        "椿弁当(おかずのみ)",
        "はなまる弁当(半ライス)",
        "はなまる弁当(おかずのみ)",
    ]

    def row_sort_key(row):
        try:
            return (PRIORITY_NAMES.index(row['mon_an']), row['mon_an'])
        except ValueError:
            return (len(PRIORITY_NAMES), row['mon_an'])

    groups['day'].sort(key=row_sort_key)
    groups['night'].sort(key=row_sort_key)
    groups['koujika'].sort(key=row_sort_key)
    groups['kyakusama'].sort(key=row_sort_key)

    summary = defaultdict(lambda: {'so_luong': 0, 'tong_tien': 0, 'gia': 0})
    for o in orders:
        s = summary[o.mon_an.ten]
        s['so_luong'] += o.so_luong
        s['gia'] = o.mon_an.gia
        s['tong_tien'] += o.so_luong * o.mon_an.gia
    summary_list = [{'ten': k, 'so_luong': v['so_luong'], 'gia': v['gia'], 'tong_tien': v['tong_tien']} for k, v in summary.items()]

    holidays = [d.strftime('%Y-%m-%d') for d in Holiday.objects.values_list('date', flat=True)]

    html = render_to_string('menu/order_kanri_pdf.html', {
        'days': days,
        'data_day': groups['day'],
        'data_night': groups['night'],
        'data_koujika': groups['koujika'],
        'data_kyakusama': groups['kyakusama'],
        'year': year,
        'month': month,
        'holidays': holidays,
        'summary_list': summary_list,
    })
    pdf_bytes = HTML(string=html).write_pdf()
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = f'filename="order_kanri_{year}_{month}.pdf"'
    return resp

# =============== DAILY DETAIL ===============
def order_detail(request, year, month, day):
    ngay = date(int(year), int(month), int(day))
    orders = Order.objects.filter(ngay_giao=ngay).select_related('mon_an')
    summary_day = {}
    summary_night = {}
    summary_koujika = {}
    summary_kyakusama = {}

    for o in orders:
        key = o.mon_an.ten
        if o.calamviec == '夜勤':
            summary_night.setdefault(key, {'ten': key, 'so_luong': 0})['so_luong'] += o.so_luong
        elif o.calamviec == '工作課':
            summary_koujika.setdefault(key, {'ten': key, 'so_luong': 0})['so_luong'] += o.so_luong
        elif o.calamviec == '客様':
            summary_kyakusama.setdefault(key, {'ten': key, 'so_luong': 0})['so_luong'] += o.so_luong
        else:
            summary_day.setdefault(key, {'ten': key, 'so_luong': 0})['so_luong'] += o.so_luong

    merged_honsya = {}
    for ten, mon_obj in summary_day.items():
        merged_honsya[ten] = {'ten': ten, 'so_luong': mon_obj['so_luong']}
    for ten, mon_obj in summary_kyakusama.items():
        merged_honsya.setdefault(ten, {'ten': ten, 'so_luong': 0})
        merged_honsya[ten]['so_luong'] += mon_obj['so_luong']

    order_groups = [
        {'title': '本社（日勤＋客様）', 'summary': merged_honsya},
        {'title': '夜勤（夜間）', 'summary': summary_night},
        {'title': '工作課', 'summary': summary_koujika},
    ]
    fax_status = FaxStatus.objects.filter(ngay=ngay).first()  # <-- thêm dòng này
    printed_flag = request.session.get(f'printed_{ngay.isoformat()}', False)  # trạng thái đã in
    return render(request, 'menu/order_detail.html', {
        'ngay': ngay,
        'order_groups': order_groups,
        'fax_status': fax_status,  # <-- thêm vào context
        'printed_flag': printed_flag,  # <-- thêm vào context
    })

# =============== AUTO YEAR COPY (GIỮ NGUYÊN) =========
def order_menu_year(request, pk):
    mon = get_object_or_404(MonAn, pk=pk)
    ma_nv = request.session.get('ma_nv')
    if not ma_nv:
        return redirect('menu:dangnhap')
    nv = NhanVien.objects.filter(ma_so=ma_nv).first()
    ten_nv = nv.ten if nv else ""

    t = today()
    last_month_order = Order.objects.filter(ma_nv=ma_nv, ngay_giao__year=t.year, ngay_giao__month=t.month).order_by('-ngay_giao').first()
    if not last_month_order:
        last_month_order = Order.objects.filter(ma_nv=ma_nv).order_by('-ngay_giao').first()
    if not last_month_order:
        messages.error(request, "まだ注文履歴がありません。")
        return redirect('menu:order_menu', pk=pk)

    mon_gan_nhat = last_month_order.mon_an
    calamviec = last_month_order.calamviec
    holidays = set(Holiday.objects.values_list('date', flat=True))
    created = 0
    for m in range(t.month, 13):
        start_d = date(t.year, m, 16)
        end_d = date(t.year + 1, 1, 15) if m == 12 else date(t.year, m + 1, 15)
        for d in working_days(start_d, end_d, holidays):
            if is_past(d, request):
                continue
            Order.objects.get_or_create(
                ma_nv=ma_nv,
                ngay_giao=d,
                defaults={
                    'ten_nv': ten_nv,
                    'mon_an': mon_gan_nhat,
                    'so_luong': 1,
                    'ghi_chu': '',
                    'calamviec': calamviec
                }
            )
            created += 1
    messages.success(request, f"{created}件の注文を1年分まとめて登録しました。")
    return redirect('menu:dangnhap')

@require_POST
def change_calamviec_multi(request):
    if not (request.user.is_authenticated and request.user.username == 'kanri'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    ma_nv = request.POST.get('ma_nv')
    calamviec = request.POST.get('calamviec')
    dates = request.POST.getlist('dates[]')
    updated = 0
    for d_str in dates:
        try:
            d = datetime.strptime(d_str, "%Y-%m-%d").date()
        except:
            continue
        # Cập nhật tất cả Order đúng mã nhân viên và ngày
        orders = Order.objects.filter(ma_nv=ma_nv, ngay_giao=d)
        for order in orders:
            order.calamviec = calamviec
            order.save()
            updated += 1
    return JsonResponse({'updated': updated})

@require_POST
def fax_mark_printed(request, year, month, day):
    """Gọi bởi JS sau khi hộp thoại in đóng; tự động đánh dấu đã gửi Fax."""
    ngay = date(int(year), int(month), int(day))
    request.session[f'printed_{ngay.isoformat()}'] = True
    fs, _ = FaxStatus.objects.get_or_create(ngay=ngay)
    if not fs.sent:
        fs.sent = True
        fs.sent_at = timezone.now()
        if request.user.is_authenticated:
            fs.user = request.user
        fs.save()
    return JsonResponse({
        'printed': True,
        'sent': True,
        'sent_at': fs.sent_at.strftime('%Y-%m-%d %H:%M') if fs.sent_at else None
    })

@require_POST
def fax_set(request, year, month, day):
    ngay = date(int(year), int(month), int(day))
    # Kiểm tra đã in chưa
    if not request.session.get(f'printed_{ngay.isoformat()}'):
        messages.error(request, "先に Ctrl+P でFAX印刷を完了してください。")
        return redirect('menu:order_detail', year=year, month=month, day=day)
    fs, _ = FaxStatus.objects.get_or_create(ngay=ngay)
    fs.sent = True
    fs.sent_at = timezone.now()
    if request.user.is_authenticated:
        fs.user = request.user
    fs.save()
    messages.success(request, "FAX送信済みを手動記録しました。")
    return redirect('menu:order_detail', year=year, month=month, day=day)

@require_POST
def fax_unset(request, year, month, day):
    ngay = date(int(year), int(month), int(day))
    fs, _ = FaxStatus.objects.get_or_create(ngay=ngay)
    fs.sent = False
    fs.sent_at = None
    fs.user = None
    fs.save()
    # Xóa cờ in nếu muốn buộc in lại:
    request.session.pop(f'printed_{ngay.isoformat()}', None)
    messages.success(request, "FAX送信記録を解除しました。")
    return redirect('menu:order_detail', year=year, month=month, day=day)

def order_kanri_excel(request):
    t = today()
    month_str = request.GET.get('month')
    if month_str:
        year, month = map(int, month_str.split('-'))
        start_date = date(year, month, 16)
        end_date = date(year + 1, 1, 15) if month == 12 else date(year, month + 1, 15)
    else:
        start_date, end_date, year, month = current_period(t)
    # Tạo danh sách ngày từ start_date đến end_date
    days = []
    d = start_date
    while d <= end_date:
        days.append(d)
        d += timedelta(days=1)
    nhanviens = NhanVien.objects.all()
    orders = Order.objects.filter(ngay_giao__gte=start_date, ngay_giao__lte=end_date).select_related('mon_an')

    # Gộp tất cả các ca làm việc cho từng nhân viên
    from collections import defaultdict
    orders_dict = defaultdict(list)
    for o in orders:
        orders_dict[o.ma_nv].append(o)

    PRIORITY_NAMES = [
        "椿弁当",
        "椿弁当(半ライス)",
        "椿弁当(おかずのみ)",
        "はなまる弁当(半ライス)",
        "はなまる弁当(おかずのみ)",
    ]

    data = []
    for nv in nhanviens:
        row = {'ma_so': nv.ma_so, 'ten': nv.ten, 'mon_an': '', 'gia': 0, 'gia2': 0, 'orders': [0]*len(days)}
        monan_counter = {}
        gia_counter = {}
        gia2_counter = {}
        for idx, d2 in enumerate(days):
            # Lấy tất cả order của nhân viên này trong ngày d2 (có thể nhiều ca)
            orders_in_day = [o for o in orders_dict.get(nv.ma_so, []) if o.ngay_giao == d2]
            if orders_in_day:
                # Gộp số lượng, lấy tên món ăn và giá của order đầu tiên (ưu tiên theo thứ tự xuất hiện)
                row['orders'][idx] = sum(o.so_luong for o in orders_in_day)
                for o in orders_in_day:
                    monan_counter[o.mon_an.ten] = monan_counter.get(o.mon_an.ten, 0) + o.so_luong
                    gia_counter[o.mon_an.ten] = int(o.mon_an.gia)
                    gia2_counter[o.mon_an.ten] = int(getattr(o.mon_an, 'gia2', 0)) or int(o.mon_an.gia)
        # Chọn món ăn có số lượng nhiều nhất trong kỳ làm đại diện
        if monan_counter:
            mon_an_max = max(monan_counter.items(), key=lambda x: x[1])[0]
            row['mon_an'] = mon_an_max
            row['gia'] = gia_counter[mon_an_max]
            row['gia2'] = gia2_counter[mon_an_max]
        # Tổng số lượng thực tế (không phải số ngày đặt)
        tong_so_luong = sum(row['orders'])
        row['total_gia'] = row['gia'] * tong_so_luong
        row['total_gia2'] = row['gia2'] * tong_so_luong
        if tong_so_luong > 0:
            data.append(row)

    def row_sort_key(row):
        try:
            return (PRIORITY_NAMES.index(row['mon_an']), row['mon_an'])
        except ValueError:
            return (len(PRIORITY_NAMES), row['mon_an'])

    data.sort(key=row_sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "注文管理"
    header = ['社員番号', '氏名', '料理名'] + [d.strftime('%Y/%m/%d') for d in days] + ['合計(マリン弁当価格)', '合計(社員価格)']
    ws.append(header)
    for r in data:
        line = [r['ma_so'], r['ten'], r['mon_an']]
        # Hiển thị số lượng đặt (có thể là 0, 1, 2,...)
        line.extend(str(qty) if qty > 0 else '' for qty in r['orders'])
        line.append(r['total_gia'])
        line.append(r['total_gia2'])
        ws.append(line)

    # Optional: Auto width
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="order_kanri_{year}_{month}.xlsx"'
    wb.save(response)
    return response

# Thêm view này:

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect

@require_POST
@csrf_protect
def copy_order(request):
    """Sao chép order từ ngày khác sang ngày được chọn"""
    import json
    from datetime import datetime
    
    order_id = request.POST.get('order_id')
    new_date = request.POST.get('new_date')
    ma_nv = request.POST.get('ma_nv')
    
    try:
        # Lấy order gốc
        original_order = Order.objects.get(id=order_id)
        
        # Kiểm tra ngày mới có order rồi không
        new_date_obj = datetime.strptime(new_date, '%Y-%m-%d').date()
        existing = Order.objects.filter(
            ma_nv=ma_nv,
            ngay_giao=new_date_obj
        )
        
        if existing.exists():
            return JsonResponse({
                'success': False,
                'error': 'この日付にはすでに注文があります'
            })
        
        # Tạo order mới
        new_order = Order(
            ma_nv=original_order.ma_nv,
            mon_an=original_order.mon_an,
            so_luong=original_order.so_luong,
            ngay_giao=new_date_obj,
            ghi_chu=original_order.ghi_chu,
            calamviec=original_order.calamviec
        )
        new_order.save()
        
        return JsonResponse({'success': True})
    
    except Order.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': '注文が見つかりません'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })