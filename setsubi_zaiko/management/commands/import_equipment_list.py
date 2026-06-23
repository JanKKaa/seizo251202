from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from setsubi_zaiko.models import EquipmentCategory, EquipmentItem


TYPE_MAP = {
    "成形機": ("EQ-SK", "injection_molding_machine"),
    "取出し機": ("EQ-TD", "takeout_robot"),
    "粉砕機": ("EQ-FS", "crusher"),
    "温調機": ("EQ-OC", "temperature_controller"),
    "乾燥機": ("EQ-KS", "dryer"),
    "真空ポンプ": ("EQ-SP", "vacuum_pump"),
    "コンプレッサー": ("EQ-KP", "compressor"),
    "金型監視カメラ": ("EQ-KC", "mold_monitor_camera"),
    "エアーベント": ("EQ-AB", "air_vent"),
    "コンベヤ": ("EQ-KB", "conveyor"),
    "洗浄機": ("EQ-SJ", "washer"),
    "脱磁機": ("EQ-DT", "demagnetizer"),
    "自動機": ("EQ-JD", "automatic_machine"),
    "混合機": ("EQ-KG", "mixer"),
    "混合機（タンブラー）": ("EQ-KG", "mixer"),
    "輸送システムローダ": ("EQ-HR", "material_loader"),
}


class Command(BaseCommand):
    help = "Import equipment masters from the production engineering equipment list workbook."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Path to 設備リスト.xlsx")
        parser.add_argument("--sheet", default="設備管理台帳")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("openpyxl is required to import 設備リスト.xlsx") from exc

        path = Path(options["xlsx_path"])
        if not path.exists():
            raise CommandError(f"Excel file not found: {path}")

        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet_name = options["sheet"]
        if sheet_name not in workbook.sheetnames:
            raise CommandError(f"Sheet not found: {sheet_name}")

        worksheet = workbook[sheet_name]
        headers = [self._text(value) for value in next(worksheet.iter_rows(min_row=4, max_row=4, values_only=True))]
        rows = [self._row_dict(headers, row) for row in worksheet.iter_rows(min_row=5, values_only=True)]
        rows = [row for row in rows if row.get("管理番号")]

        created = 0
        updated = 0
        skipped = 0
        dry_run = options["dry_run"]

        with transaction.atomic():
            for row in rows:
                category_code, equipment_type = TYPE_MAP.get(row.get("種類"), ("EQUIPMENT-LEDGER", "other"))
                category = EquipmentCategory.objects.filter(code=category_code).first()
                if category is None:
                    skipped += 1
                    self.stderr.write(f"Skip {row.get('管理番号')}: category {category_code} not found")
                    continue

                defaults = self._defaults(row, category, equipment_type)
                _, was_created = EquipmentItem.objects.update_or_create(
                    code=row["管理番号"],
                    defaults=defaults,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

            if dry_run:
                transaction.set_rollback(True)

        suffix = " (dry-run, rolled back)" if dry_run else ""
        self.stdout.write(self.style.SUCCESS(f"Equipment import: created={created}, updated={updated}, skipped={skipped}{suffix}"))

    def _defaults(self, row, category, equipment_type):
        equipment_kind = row.get("種類") or ""
        model_no = row.get("型式") or ""
        name = f"{equipment_kind} {model_no}".strip()[:160] or row["管理番号"]
        note_parts = [
            ("台帳No.", row.get("No.")),
            ("番号順", row.get("番号順")),
            ("製造年月", row.get("製造年月")),
            ("電源", row.get("電源（電圧）")),
            ("エアー", row.get("エアー")),
            ("材料名", row.get("材料名")),
            ("製品名", row.get("製品名")),
            ("備考", row.get("備考")),
        ]
        note = "\n".join(f"{label}: {value}" for label, value in note_parts if value)
        return {
            "name": name,
            "category": category,
            "equipment_type": equipment_type,
            "item_kind": EquipmentItem.KIND_EQUIPMENT,
            "equipment_group_name": equipment_kind[:120],
            "equipment_series_name": model_no[:120],
            "maker": row.get("メーカー") or "",
            "model_no": model_no[:100],
            "serial_no": (row.get("製造番号") or "")[:100],
            "applicable_machine_no": row.get("成形号機") or "",
            "location": "設備管理台帳",
            "department": "生産技術課",
            "status": "in_use",
            "current_quantity": Decimal("1"),
            "unit": "式",
            "quality_rank": "C",
            "note": note,
        }

    def _row_dict(self, headers, row):
        result = {}
        for index, header in enumerate(headers):
            if header:
                result[header] = self._text(row[index] if index < len(row) else None)
        return result

    def _text(self, value):
        if value is None:
            return ""
        return str(value).strip()
